"""Bulk import prospects from CSV/Excel files."""
import csv
import io
import logging
from typing import List, Tuple
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class BulkImportService:
    REQUIRED_COLUMNS = {"full_name", "company", "email"}
    VALID_COLUMNS = {
        "full_name", "company", "email", "title", "linkedin_url",
        "instagram_handle", "company_domain", "twitter_handle", "phone", "location"
    }

    def parse_file(self, file_content: bytes, filename: str) -> Tuple[List[dict], List[str]]:
        """Parse CSV or Excel file. Returns (rows, errors)."""
        errors = []
        rows = []

        try:
            if filename.lower().endswith(".xlsx"):
                rows = self._parse_excel(file_content)
            elif filename.lower().endswith(".csv"):
                rows = self._parse_csv(file_content)
            else:
                errors.append(f"Unsupported file format. Use .csv or .xlsx")
        except Exception as e:
            errors.append(f"Failed to parse file: {str(e)}")

        return rows, errors

    def _parse_excel(self, content: bytes) -> List[dict]:
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        headers = [h.value for h in next(ws.iter_rows(max_row=1))]
        headers = [h.strip().lower().replace(" ", "_") if h else None for h in headers]

        valid_headers = []
        for h in headers:
            if h and h in self.VALID_COLUMNS:
                valid_headers.append(h)

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            row_dict = {}
            for i, h in enumerate(valid_headers):
                if i < len(row):
                    row_dict[h] = row[i]
            if row_dict.get("full_name") or row_dict.get("company"):
                rows.append(row_dict)

        wb.close()
        return rows

    def _parse_csv(self, content: bytes) -> List[dict]:
        decoded = content.decode("utf-8", errors="replace")
        reader = csv.DictReader(io.StringIO(decoded))
        rows = []
        for row in reader:
            cleaned = {
                k.strip().lower().replace(" ", "_"): v.strip() if v else None
                for k, v in row.items() if k
            }
            if cleaned.get("full_name") or cleaned.get("company"):
                rows.append(cleaned)
        return rows

    def validate_row(self, row: dict, row_num: int) -> Tuple[bool, str]:
        """Validate a single row. Returns (valid, error_message)."""
        if not row.get("email"):
            return False, f"Row {row_num}: missing email"
        if "@" not in row.get("email", ""):
            return False, f"Row {row_num}: invalid email format"
        if not row.get("full_name") and not row.get("company"):
            return False, f"Row {row_num}: missing full_name or company"
        return True, ""

    def build_prospect_payload(self, row: dict, user_id: str) -> dict:
        """Build a prospect record from a validated row."""
        full_name = row.get("full_name", "") or ""
        first_name = full_name.split()[0] if full_name else ""
        last_name = " ".join(full_name.split()[1:]) if len(full_name.split()) > 1 else ""

        linkedin_url = row.get("linkedin_url") or None
        if linkedin_url and not linkedin_url.startswith(("http://", "https://")):
            linkedin_url = f"https://{linkedin_url}"

        instagram = row.get("instagram_handle") or None
        if instagram and instagram.startswith("@"):
            instagram = instagram[1:]

        twitter = row.get("twitter_handle") or None
        if twitter and twitter.startswith("@"):
            twitter = twitter[1:]

        return {
            "user_id": user_id,
            "full_name": full_name,
            "first_name": first_name,
            "last_name": last_name,
            "email": row.get("email"),
            "company": row.get("company") or "",
            "title": row.get("title") or "",
            "linkedin_url": linkedin_url,
            "instagram_handle": instagram,
            "company_domain": row.get("company_domain") or "",
            "twitter_handle": twitter,
            "phone": row.get("phone") or "",
            "location": row.get("location") or "",
            "source": "bulk_import",
        }
